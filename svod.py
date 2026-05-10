import pandas as pd

import os
from str_pptx import generate_table, keys_table_svod_summer, keys_table_svod
from python_pptx_text_replacer import TextReplacer
import locale
from pptx_functions import (
    remove_slides_tinao,
    convert_pptx_to_pdf,
    pdf_to_png,
    runs_from_pptx_svod,
)
from datetime import datetime, timedelta
from openpyxl.drawing.image import Image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side
import comtypes.client


order = [
    "ЦАО",
    "САО",
    "СВАО",
    "ВАО",
    "ЮВАО",
    "ЮАО",
    "ЮЗАО",
    "ЗАО",
    "СЗАО",
    "ЗелАО",
    "ТиНАО",
]

allert = []


def fint(x):
    locale.setlocale(locale.LC_ALL, "ru_RU.UTF-8")
    return locale._format("%d", x, grouping=True)


def format_value(value):
    """
    Безопасное форматирование значений для вставки в pptx.
    """
    if pd.isna(value):
        return ""

    if isinstance(value, (int, float)) and float(value).is_integer():
        return fint(int(value))

    return str(value)


def normalize_input_data(df_ais, df_edc, ais_event_name):
    """
    Нормализация входных данных.
    """

    df_ais = df_ais.copy()
    df_edc = df_edc.copy()

    if "Округ" in df_ais.columns:
        df_ais["Округ"] = df_ais["Округ"].fillna("").astype(str).str.strip()
        df_ais["Округ"] = df_ais["Округ"].replace(
            {
                "НАО": "ТиНАО",
                "ТАО": "ТиНАО",
            }
        )

    if "Округ" in df_edc.columns:
        df_edc["Округ"] = df_edc["Округ"].fillna("").astype(str).str.strip()
        df_edc["Округ"] = df_edc["Округ"].replace(
            {
                "НАО": "ТиНАО",
                "ТАО": "ТиНАО",
            }
        )

    if "Район" in df_ais.columns:
        df_ais["Район"] = df_ais["Район"].fillna("").astype(str).str.strip()

    if "Район" in df_edc.columns:
        df_edc["Район"] = df_edc["Район"].fillna("").astype(str).str.strip()

    if ais_event_name in df_ais.columns:
        df_ais[ais_event_name] = (
            df_ais[ais_event_name]
            .fillna("")
            .astype(str)
            .str.strip()
        )

    return df_ais, df_edc


def check_pptx_mapping(label, flat_keys, flat_values):
    """
    Диагностика перед вставкой в pptx.
    """
    global allert

    print(
        label,
        "keys:",
        len(flat_keys),
        "values:",
        len(flat_values),
        "unique keys:",
        len(set(flat_keys)),
    )

    if len(flat_keys) != len(flat_values):
        allert.append(
            f"Ошибка PPTX [{label}]: ключей {len(flat_keys)}, "
            f"значений {len(flat_values)}. Данные могут вставиться неправильно."
        )

    if len(flat_keys) != len(set(flat_keys)):
        allert.append(
            f"Ошибка PPTX [{label}]: есть повторяющиеся ключи. "
            f"Часть значений может быть перезатёрта."
        )


def make_header_values_for_district_slide(streets, keys_count):
    """
    Формирует значения для заголовков районов.

    В макете заголовки районов сделаны в 2 строки:
    - первая строка: районы + Итого
    - вторая строка: пустые значения

    Поэтому количество значений должно совпадать с количеством ключей.
    """

    first_row = list(streets) + ["Итого"]
    second_row = [""] * len(first_row)

    values = first_row + second_row

    if len(values) < keys_count:
        values += [""] * (keys_count - len(values))

    if len(values) > keys_count:
        values = values[:keys_count]

    return values


def insert_images_to_excel(image_paths, excel_files):
    print(image_paths, excel_files)

    """
    Вставляет изображения в Excel, создаёт новый лист с именем файла
    и ставит его первым.
    """

    if len(image_paths) < len(excel_files) + 1:
        print("Ошибка: Слишком мало слайдов для всех файлов Excel")
        return

    for i, excel_path in enumerate(excel_files):
        if not os.path.exists(excel_path):
            print(f"Файл не найден, создаю новый: {excel_path}")
            wb = Workbook()
        else:
            wb = load_workbook(excel_path)

        img_path = image_paths[i + 1]
        sheet_name = order[i]

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(sheet_name)

        img = Image(img_path)
        ws.add_image(img, "B2")

        wb.move_sheet(ws, offset=-len(wb.sheetnames) + 1)

        wb.save(excel_path)
        wb.close()

        print(f"Обновлён файл Excel: {excel_path}, вставлен {img_path}")

    comtypes.CoInitialize()

    excel_app = comtypes.client.CreateObject("Excel.Application")
    excel_app.Visible = False
    excel_app.DisplayAlerts = False

    try:
        for excel_path in excel_files:
            abs_path = os.path.abspath(excel_path)
            wb_com = excel_app.Workbooks.Open(abs_path)

            ws_com = wb_com.Worksheets(1)

            for shape in ws_com.Shapes:
                if shape.Type == 13:
                    shape.Left = shape.Left - 15

            wb_com.Save()
            wb_com.Close()

            print(
                f"COM: смещены изображения в файле {excel_path} "
                f"на 15px влево (первый лист)"
            )

    finally:
        excel_app.Quit()
        comtypes.CoUninitialize()


def make_svod_presentation(ais_file, edc_file, date, morning, summer):
    global allert
    allert = []

    order = [
        "ЦАО",
        "САО",
        "СВАО",
        "ВАО",
        "ЮВАО",
        "ЮАО",
        "ЮЗАО",
        "ЗАО",
        "СЗАО",
        "ЗелАО",
        "ТиНАО",
        'ГБУ "АВД"',
        "Иные",
        "Общий итог",
    ]

    presentation_maket = "makets/presentation/svod_presentation.pptx"

    if summer:
        presentation_maket = "makets/presentation/svod_presentation-sumr.pptx"

    replacer = TextReplacer(
        presentation_maket,
        slides="",
        tables=True,
        charts=True,
        textframes=True,
        quiet=True,
    )

    time = "17:00"
    date_text = f"{date} на {time}"

    tmp_files_path = os.path.join(
        "",
        "reports",
        f"{date_text}",
        "tmp_files",
    ).replace(":", ".")

    path_os = f"reports/{date_text}".replace(":", ".")

    os.makedirs(f"{path_os}/Сводка", exist_ok=True)
    os.makedirs(tmp_files_path, exist_ok=True)

    f_time = "17:00"
    f_date = date

    if morning:
        date_obj = datetime.strptime(date, "%d.%m.%Y")
        date_obj -= timedelta(days=1)
        f_date = str(date_obj.strftime("%d.%m.%Y"))

    date_svod_text = f"с {f_time} {f_date} по {time} {date}".replace(":", ".")

    xlsx_folder = f"{path_os}/Сводка"

    xlsx_files = [
        os.path.join(xlsx_folder, f"{name} {date_svod_text}.xlsx")
        for name in order
    ]

    xlsx_files.pop()
    xlsx_files.pop()
    xlsx_files.pop()

    ais_event_name = "Наименование события"

    df_ais = pd.read_excel(ais_file)
    df_edc = pd.read_excel(edc_file)

    df_ais, df_edc = normalize_input_data(
        df_ais=df_ais,
        df_edc=df_edc,
        ais_event_name=ais_event_name,
    )

    # =========================
    # Общая сводная таблица
    # =========================

    pivot_table = pd.pivot_table(
        df_ais,
        index=ais_event_name,
        columns="Округ",
        aggfunc="size",
        fill_value=0,
    )

    pivot_table["Итого по строке"] = pivot_table.sum(axis=1)

    ordered_columns = [
        col for col in order if col in pivot_table.columns
    ] + ["Итого по строке"]

    pivot_table = pivot_table[ordered_columns]

    original_pivot_table = pivot_table.copy()

    pivot_table_sorted = pivot_table.sort_values(
        by="Итого по строке",
        ascending=False,
    )

    top_10 = pivot_table_sorted.iloc[:10]

    if not summer:
        edc_summary = df_edc["Округ"].value_counts().reindex(order, fill_value=0)
        edc_summary["Итого по строке"] = edc_summary.sum()
        edc_summary.name = "Пуск отопления"

        edc_row = pd.DataFrame(edc_summary).T

        pivot_table_with_heating = pd.concat([top_10, edc_row])
    else:
        pivot_table_with_heating = top_10.copy()

    pivot_table_with_heating = pivot_table_with_heating.sort_values(
        by="Итого по строке",
        ascending=False,
    )

    remaining_rows = original_pivot_table.drop(
        index=pivot_table_with_heating.index,
        errors="ignore",
    )

    other_row = remaining_rows.sum(axis=0)
    other_row.name = "Иные"

    pivot_table_with_heating = pd.concat(
        [
            pivot_table_with_heating,
            pd.DataFrame([other_row]),
        ]
    )

    column_sums = pivot_table_with_heating.sum(axis=0)
    column_sums.name = "Итого по столбцу"

    pivot_table_with_heating = pd.concat(
        [
            pivot_table_with_heating,
            pd.DataFrame([column_sums]),
        ]
    )

    final_table = pivot_table_with_heating[ordered_columns]

    result_svod_path = f"{tmp_files_path}/Результаты по своду.xlsx"

    with pd.ExcelWriter(
        result_svod_path,
        mode="w",
        engine="openpyxl",
    ) as writer:
        final_table.to_excel(writer, sheet_name="Сводная таблица")

    top_10_events = [
        event
        for event in final_table.index
        if event not in ["Иные", "Итого по столбцу", "Пуск отопления"]
    ]

    tinao_len = 0

    # =========================
    # Excel-файлы и районные слайды
    # =========================

    for i, district in enumerate(order):
        if district == "Общий итог":
            continue

        df_district_ais = df_ais[
            (df_ais["Округ"] == district)
            & (df_ais[ais_event_name].isin(top_10_events))
        ]

        df_district_ais_spec_xl = df_ais[
            df_ais["Округ"] == district
        ]

        df_district_ais_xl = df_district_ais_spec_xl.copy()

        df_district_edc = df_edc[
            df_edc["Округ"] == district
        ]

        selected_columns = [
            "№ во внешней системе",
            "№ в системе",
            "Наименование события",
            "Система",
            "Ответственный",
            "Адрес объекта",
            "Район",
            "Округ",
            "Дата создания во внешней системе",
        ]

        existing_columns = [
            col for col in selected_columns
            if col in df_district_ais_xl.columns
        ]

        df_district_ais_for_xl = df_district_ais_xl[existing_columns].copy()

        df_district_ais_for_xl.loc[:, "Статус"] = ""
        df_district_ais_for_xl.loc[:, "Примечание"] = ""

        district_clear = district.replace('"', "")

        xlsx_path = f"{path_os}/Сводка/{district_clear} {date_svod_text}.xlsx"

        with pd.ExcelWriter(
            xlsx_path,
            mode="w",
            engine="openpyxl",
        ) as writer:
            df_district_ais_for_xl.to_excel(
                writer,
                sheet_name="АИС ЦУ КГХ",
                index=False,
            )

        wb = load_workbook(xlsx_path)
        ws = wb["АИС ЦУ КГХ"]

        ws.insert_rows(1)
        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=11,
        )

        header_cell = ws.cell(row=1, column=1, value=district)

        column_width = 16

        for col in range(1, 12):
            if col in [5, 6]:
                ws.column_dimensions[
                    ws.cell(row=2, column=col).column_letter
                ].width = (column_width * 2) + 6
            else:
                ws.column_dimensions[
                    ws.cell(row=2, column=col).column_letter
                ].width = column_width

        row_height = 56

        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = row_height

            if row == 1:
                ws.row_dimensions[row].height = 32

        border_style = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=11,
        ):
            for cell in row:
                cell.border = border_style
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="center",
                )

        header_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        wb.save(xlsx_path)
        wb.close()

        print(f"{district}")

        if not summer:
            with pd.ExcelWriter(
                xlsx_path,
                mode="a",
                engine="openpyxl",
                if_sheet_exists="replace",
            ) as writer:
                df_district_edc.to_excel(
                    writer,
                    sheet_name="ЕДЦ",
                    index=False,
                )

        if district == 'ГБУ "АВД"' or district == "Иные":
            print(f"{district} не рисует слайд")
            continue

        # =========================
        # Районная таблица для слайда
        # =========================

        pivot_table_district = pd.pivot_table(
            df_district_ais,
            index=ais_event_name,
            columns="Район",
            aggfunc="size",
            fill_value=0,
        )

        pivot_table_district = pivot_table_district.reindex(
            top_10_events,
            fill_value=0,
        )

        df_district_ais_other = df_ais.loc[
            (df_ais["Округ"] == district)
            & (~df_ais[ais_event_name].isin(top_10_events)),
            ["Район"],
        ].copy()

        other_summary_district = df_district_ais_other["Район"].value_counts()
        other_summary_district.name = "Иные"

        other_row_district = pd.DataFrame(other_summary_district).T

        if not summer:
            edc_summary_district = df_district_edc["Район"].value_counts()
            edc_summary_district.name = "Пуск отопления"

            edc_row_district = pd.DataFrame(edc_summary_district).T

            all_districts = list(
                dict.fromkeys(
                    list(pivot_table_district.columns)
                    + list(edc_row_district.columns)
                    + list(other_row_district.columns)
                )
            )
        else:
            edc_row_district = None

            all_districts = list(
                dict.fromkeys(
                    list(pivot_table_district.columns)
                    + list(other_row_district.columns)
                )
            )

        all_districts = [
            x for x in all_districts
            if str(x).strip() not in ["", "nan", "None"]
        ]

        pivot_table_district = pivot_table_district.reindex(
            columns=all_districts,
            fill_value=0,
        )

        other_row_district = other_row_district.reindex(
            columns=all_districts,
            fill_value=0,
        )

        if not summer:
            edc_row_district = edc_row_district.reindex(
                columns=all_districts,
                fill_value=0,
            )

            if "Пуск отопления" not in pivot_table_district.index:
                pivot_table_district.loc["Пуск отопления"] = 0

            pivot_table_district.loc["Пуск отопления", all_districts] = (
                pivot_table_district.loc["Пуск отопления", all_districts]
                + edc_row_district.loc["Пуск отопления", all_districts]
            )

        if "Иные" not in pivot_table_district.index:
            pivot_table_district.loc["Иные"] = 0

        pivot_table_district.loc["Иные", all_districts] = (
            pivot_table_district.loc["Иные", all_districts]
            + other_row_district.loc["Иные", all_districts]
        )

        pivot_table_district["Итого по строке"] = (
            pivot_table_district[all_districts].sum(axis=1)
        )

        column_sums_district = pivot_table_district.sum(axis=0)
        column_sums_district.name = "Итого по столбцу"

        pivot_table_with_heating_district = pd.concat(
            [
                pivot_table_district,
                pd.DataFrame([column_sums_district]),
            ]
        )

        num_columns = pivot_table_district.shape[1]

        letters = [
            "c",
            "s",
            "sv",
            "v",
            "yv",
            "y",
            "yz",
            "z",
            "sz",
            "ze",
            "tin",
        ]

        row_len = [
            11,
            17,
            18,
            17,
            13,
            17,
            13,
            13,
            9,
            6,
            6,
        ]

        if district == "ТиНАО":
            letter = "tin"
            tinao_len = num_columns

            if num_columns > 13:
                allert.append("Внимание! У ТИНАО больше 12 районов!!!")
        else:
            letter = letters[i]

            if num_columns < row_len[i]:
                allert.append(f"Внимание! У {district} меньше районов!")
            elif num_columns > row_len[i]:
                allert.append(f"Внимание! У {district} больше районов!")

        if not summer:
            keys_district = generate_table(letter, num_columns, 13)
        else:
            keys_district = generate_table(letter, num_columns, 12, 1)

        flat_keys = [
            key
            for row in keys_district
            for key in row
        ]

        flat_values = pivot_table_with_heating_district.to_numpy().flatten()

        check_pptx_mapping(
            f"районная таблица {district}",
            flat_keys,
            flat_values,
        )

        table_dict = dict(zip(flat_keys, flat_values))

        table_list = [
            (
                key,
                format_value(value),
            )
            for key, value in table_dict.items()
        ]

        # =========================
        # Заголовки районов для слайда
        # =========================

        streets = list(pivot_table_with_heating_district.columns[:-1])

        streets_key = generate_table(
            f"r{letters[i]}",
            len(streets) + 1,
            2,
        )

        s_flat_keys = [
            key
            for row in streets_key
            for key in row
        ]

        s_flat_values = make_header_values_for_district_slide(
            streets=streets,
            keys_count=len(s_flat_keys),
        )

        check_pptx_mapping(
            f"заголовки районов {district}",
            s_flat_keys,
            s_flat_values,
        )

        s_dict = dict(zip(s_flat_keys, s_flat_values))

        s_table_list = [
            (
                key,
                format_value(value),
            )
            for key, value in s_dict.items()
        ]

        replacer.replace_text(table_list)
        replacer.replace_text(s_table_list)

        with pd.ExcelWriter(
            result_svod_path,
            mode="a",
            engine="openpyxl",
            if_sheet_exists="replace",
        ) as writer:
            pivot_table_with_heating_district.to_excel(
                writer,
                sheet_name=district,
            )

    # =========================
    # Заполнение общей таблицы в pptx
    # =========================

    if summer:
        flat_keys = [
            key
            for row in keys_table_svod_summer
            for key in row
        ]
    else:
        flat_keys = [
            key
            for row in keys_table_svod
            for key in row
        ]

    final_table = final_table.reset_index()

    flat_values = final_table.to_numpy().flatten()

    check_pptx_mapping(
        "главная таблица",
        flat_keys,
        flat_values,
    )

    table_dict = dict(zip(flat_keys, flat_values))

    print(table_dict)

    table_list = [
        (
            key,
            format_value(value),
        )
        for key, value in table_dict.items()
    ]

    replacer.replace_text(table_list)

    replacer.replace_text(
        [
            ("*dateperiod*", date_svod_text),
        ]
    )

    # =========================
    # Сохранение презентации
    # =========================

    file_path_save = (
        f"reports/{date_text}/Сводка/Обращения граждан {date_svod_text}.pptx"
        .replace(":", ".")
    )

    raw_pptx_path = f"{tmp_files_path}/Свод_без_обработки.pptx"
    colored_pptx_path = f"{tmp_files_path}/Свод_без_обработки_крашенный.pptx"

    replacer.write_presentation_to_file(raw_pptx_path)

    file_path_save_pdf = file_path_save.replace("pptx", "pdf")

    runs_from_pptx_svod(raw_pptx_path).save(colored_pptx_path)

    remove_slides_tinao(
        colored_pptx_path,
        tinao_len,
    ).save(file_path_save)

    converted = None

    try:
        converted = convert_pptx_to_pdf(
            file_path_save,
            file_path_save_pdf,
        )
    except Exception as e:
        allert.append(f"{e}")

    if not converted:
        allert.append("Не удалось конвертировать в PDF")
        return allert

    output_folder_png = f"{tmp_files_path}/img"

    slides = pdf_to_png(
        file_path_save_pdf,
        output_folder_png,
    )

    insert_images_to_excel(
        slides,
        xlsx_files,
    )

    return allert